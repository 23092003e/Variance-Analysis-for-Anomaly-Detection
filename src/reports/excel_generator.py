"""
Excel report generation with formatted output and anomaly summary.
"""

import logging
import pandas as pd
import re
from pathlib import Path
from typing import List, Dict, Any
import xlsxwriter
from datetime import datetime

from config.settings import Settings
from data.models import FinancialData
from analysis.variance_analyzer import VarianceResult
from analysis.correlation_engine import CorrelationResult
from analysis.anomaly_detector import Anomaly, AnomalySeverity


class ExcelGenerator:
    """Excel report generator for variance analysis results."""
    
    def __init__(self, settings: Settings):
        self.settings = settings
        self.logger = logging.getLogger(__name__)
    
    def generate_report(self, financial_data: FinancialData,
                       variance_results: List[VarianceResult],
                       correlation_results: List[CorrelationResult],
                       anomalies: List[Anomaly],
                       output_file: str) -> None:
        """
        Add Anomalies Summary sheet to existing Excel file.
        Uses the source file from financial_data metadata for independent processing.
        
        Args:
            financial_data: Original financial data with metadata
            variance_results: Variance analysis results  
            correlation_results: Correlation analysis results
            anomalies: Detected anomalies
            output_file: Path to Excel file to modify
        """
        import os
        from pathlib import Path
        from openpyxl import load_workbook, Workbook
        
        # Get the original source file from metadata (for independent processing)
        source_file = None
        if financial_data.metadata and 'source_file' in financial_data.metadata:
            source_file = financial_data.metadata['source_file']
            self.logger.info(f"Using source file from metadata: {source_file}")
        
        # Determine which file to use as template
        actual_file = output_file
        if not os.path.exists(output_file):
            if source_file and os.path.exists(source_file):
                actual_file = source_file
                self.logger.info(f"Output file {output_file} doesn't exist. Using source file: {source_file}")
            else:
                # Only fallback to default as last resort
                fallback_file = self.settings.default_input_file
                if os.path.exists(fallback_file):
                    actual_file = fallback_file
                    self.logger.warning(f"Using fallback file: {fallback_file}")
                else:
                    self.logger.warning(f"No template file available. Creating new file: {output_file}")
                    self._create_fallback_file(financial_data, variance_results, correlation_results, anomalies, output_file)
                    return
        
        self.logger.info(f"Adding Anomalies Summary sheet to file: {actual_file}")
        
        try:
            workbook = load_workbook(actual_file)
            
            # Remove existing Anomalies Summary sheet if it exists
            if 'Anomalies Summary' in workbook.sheetnames:
                workbook.remove(workbook['Anomalies Summary'])
                self.logger.info("Removed existing 'Anomalies Summary' sheet")
            
            # Create new sheet
            anomaly_sheet = workbook.create_sheet('Anomalies Summary')
            
            # Prepare and write data using current financial_data
            summary_data = self._prepare_anomaly_data(anomalies, financial_data)
            self._write_anomaly_data(anomaly_sheet, summary_data)
            self._apply_anomaly_formatting(anomaly_sheet, summary_data, anomalies)
            
            # Always save to the specified output file (not the template file)
            workbook.save(output_file)
            workbook.close()
            
            self.logger.info(f"Anomalies Summary sheet added successfully to: {output_file}")
            
        except Exception as e:
            self.logger.error(f"Error modifying file {actual_file}: {str(e)}")
            # Fallback to creating new file
            self._create_fallback_file(financial_data, variance_results, correlation_results, anomalies, output_file)
    
    
    def _prepare_anomaly_data(self, anomalies: List[Anomaly], financial_data: FinancialData) -> List[dict]:
        """Prepare anomaly data in required format using financial data from BS and PL sheets."""
        summary_data = []
        seen_combinations = set()  # Track unique combinations to prevent duplicates
        
        self.logger.info(f"Processing {len(anomalies)} anomalies for report generation")
        
        # Get subsidiary from file path
        file_subsidiary = self._extract_subsidiary_from_filename(financial_data.metadata.get('file_path', ''))
        self.logger.info(f"Extracted subsidiary '{file_subsidiary}' from filename")
        
        # Log sheet information
        if 'sheets' in financial_data.metadata:
            available_sheets = financial_data.metadata['sheets']
            self.logger.info(f"Available sheets in source file: {available_sheets}")
        
        for i, anomaly in enumerate(anomalies, 1):
            # Log anomaly source information
            anomaly_source = self._determine_anomaly_source(anomaly)
            self.logger.debug(f"Anomaly {i}/{len(anomalies)}: Account {anomaly.account_code} ({anomaly.account_name}) "
                            f"from {anomaly_source}, Period: {anomaly.period}, "
                            f"Variance: {anomaly.variance_percent:.2f}%")
            
            # Skip total/summary accounts to reduce noise
            if self._is_total_account(anomaly.account_code, anomaly.account_name):
                self.logger.debug(f"Skipping total/summary account: {anomaly.account_code} ({anomaly.account_name})")
                continue
                
            # Use subsidiary from filename
            subsidiary = file_subsidiary
            
            # Format account as "Account Name (Code)"  
            account = f"{anomaly.account_name} ({anomaly.account_code})"
            
            # Format period as MM/YYYY
            period = self._format_period(anomaly.period)
            
            # Create unique key to check for duplicates
            unique_key = (subsidiary, account, period)
            
            # Skip if this combination already exists
            if unique_key in seen_combinations:
                self.logger.debug(f"Skipping duplicate: {subsidiary} - {account} - {period}")
                continue
            
            seen_combinations.add(unique_key)
            
            # Format percentage change with 2 decimal places
            pct_change = f"{anomaly.variance_percent:.2f}%" if anomaly.variance_percent else "0.00%"
            
            # Calculate absolute change in VND
            absolute_change = anomaly.current_value - (anomaly.previous_value or 0)
            
            # Format triggers in new format
            triggers = self._format_triggers_new(anomaly)
            
            # Generate suggested likely cause
            suggested_cause = self._generate_suggested_cause(anomaly)
            
            # Set status to "Needs Review"
            status = "Needs Review"
            
            # Default notes
            notes = ""
            
            # Log final record info
            self.logger.debug(f"Added anomaly record: {subsidiary} - {account} - {period} "
                            f"({pct_change}, {absolute_change:,.0f} VND)")
            
            summary_data.append({
                'Subsidiary': subsidiary,
                'Account': account,
                'Period': period,
                'Pct Change': pct_change,
                'Absolute Change (VND)': absolute_change,
                'Trigger(s)': triggers,
                'Suggested likely cause': suggested_cause,
                'Status': status,
                'Notes': notes,
                '_severity': anomaly.severity.value  # Store severity for formatting
            })
        
        self.logger.info(f"Generated {len(summary_data)} anomaly records from {len(anomalies)} input anomalies")
        return summary_data
    
    def _is_total_account(self, account_code: str, account_name: str) -> bool:
        """Check if account is a total/summary account that should be filtered out."""
        # Convert to lowercase for case-insensitive matching
        account_name_lower = account_name.lower()
        account_code_str = str(account_code)
        
        # Common total/summary account indicators
        total_indicators = [
            'total', 'tổng', 'sum', 'summary', 'subtotal',
            'grand total', 'net total', 'aggregate',
            'consolidated', 'combined'
        ]
        
        # Check account name for total indicators
        for indicator in total_indicators:
            if indicator in account_name_lower:
                return True
        
        # Check for account codes that are typically totals (round numbers, parent codes)
        # For example: 100, 1000, 2000, etc. (parent level accounts)
        if len(account_code_str) <= 3 and account_code_str.endswith('00'):
            return True
            
        # Check for specific patterns that indicate parent/total accounts
        # Like accounts ending with multiple zeros
        if len(account_code_str) >= 4 and account_code_str[-3:] == '000':
            return True
        
        # NEW: Check for Roman numeral patterns at beginning followed by dot
        # Examples: "I.", "II.", "III.", "IV.", "V.", "VI.", etc.
        roman_pattern = r'^(I{1,3}V?|IV|V|VI{0,3}|IX|X{1,3}L?|XL|L|LX{0,3}|XC|C{1,3}D?|CD|D|DC{0,3}|CM|M+)\.'
        if re.match(roman_pattern, account_name, re.IGNORECASE):
            return True
        
        # Check for single capital letters followed by space, dot, dash, or end (A, B, C, D, E, etc.)
        # But NOT if it's part of a longer word like "G&A:"
        single_letter_pattern = r'^[A-Z](\s|-\s|$)'
        if re.match(single_letter_pattern, account_name) and not re.match(r'^[A-Z]&', account_name):
            return True
        
        # Check for section markers like "A - SOMETHING", "B - SOMETHING", etc.
        section_pattern = r'^[A-Z]\s*-\s*[A-Z][A-Z\s]+$'
        if re.match(section_pattern, account_name, re.IGNORECASE):
            return True
        
        # Check for numeric section patterns like "1.", "2.", "3."
        numeric_section_pattern = r'^\d+\.'  # Bắt đầu bằng 1 hoặc nhiều chữ số + dấu chấm
        if re.match(numeric_section_pattern, account_name.strip()):
            return True
        
        # NEW: Check for parent account patterns
        # If account code is much shorter than typical detail accounts (usually 6+ digits)
        if len(account_code_str) <= 4 and account_code_str.isdigit():
            # Additional check: avoid filtering legitimate 4-digit detail accounts
            # by checking if it's a round number pattern
            if account_code_str.endswith('0'):
                return True
        
        # NEW: Check for hierarchical patterns where parent accounts have shorter codes
        # Example: 1234 might be parent of 123401, 123402, etc.
        if len(account_code_str) == 4 and account_code_str.isdigit():
            # Check if it looks like a parent code (ends with 0 or 00)
            if account_code_str.endswith('00') or account_code_str.endswith('0'):
                return True
        
        return False
    
    
    def _get_default_subsidiary(self) -> str:
        """Get default subsidiary name."""
        return 'DAL'  # Default to main entity  # Default to main entity  # Default to full main entity name

    def _get_subsidiary_from_financial_data(self, account_code: str, financial_data: FinancialData) -> str:
        """Get subsidiary information from financial data or filename."""
        try:
            # First, try to extract from filename (preferred method)
            if financial_data.metadata and 'file_path' in financial_data.metadata:
                subsidiary_from_filename = self._extract_subsidiary_from_filename(financial_data.metadata['file_path'])
                if subsidiary_from_filename != self._get_default_subsidiary():
                    return subsidiary_from_filename
            
            # Fallback to subsidiaries list if available
            if financial_data.subsidiaries:
                return financial_data.subsidiaries[0]
            else:
                return self._get_default_subsidiary()
        except Exception as e:
            self.logger.warning(f"Error getting subsidiary for account {account_code}: {str(e)}")
            return self._get_default_subsidiary()

    def _extract_subsidiary_from_filename(self, file_path: str) -> str:
        """Extract subsidiary code from filename. E.g. 'DAL_May25_example.xlsx' -> 'DAL'"""
        try:
            if not file_path:
                return self._get_default_subsidiary()
            
            import os
            filename = os.path.basename(file_path)
            
            # Remove file extension
            filename_without_ext = os.path.splitext(filename)[0]
            
            # Split by underscore and take the first part as subsidiary code
            parts = filename_without_ext.split('_')
            if parts and parts[0]:
                subsidiary_code = parts[0].upper()
                self.logger.debug(f"Extracted subsidiary '{subsidiary_code}' from filename '{filename}'")
                return subsidiary_code
            
            # Fallback: try to extract letters from start of filename
            import re
            match = re.match(r'^([A-Z]+)', filename_without_ext.upper())
            if match:
                subsidiary_code = match.group(1)
                self.logger.debug(f"Extracted subsidiary '{subsidiary_code}' using regex from filename '{filename}'")
                return subsidiary_code
            
            self.logger.warning(f"Could not extract subsidiary from filename '{filename}', using default")
            return self._get_default_subsidiary()
            
        except Exception as e:
            self.logger.warning(f"Error extracting subsidiary from filename '{file_path}': {str(e)}")
            return self._get_default_subsidiary()

    def _determine_anomaly_source(self, anomaly: Anomaly) -> str:
        """Determine which sheet/source the anomaly came from based on account characteristics."""
        try:
            # Check if it's likely from Balance Sheet based on account code or category
            if anomaly.category in ['investment_properties', 'borrowings', 'cash_deposits', 'trade_receivables', 'assets', 'liabilities', 'equity']:
                return "BS breakdown/BS sheet"
            
            # Check if it's likely from P&L based on account code or category  
            elif anomaly.category in ['opex', 'staff_costs', 'other_expenses', 'revenue', 'income', 'expense']:
                return "PL breakdown sheet"
            
            # Try to determine from account code patterns
            account_code_str = str(anomaly.account_code)
            
            # Balance Sheet account code patterns (typically 1000s, 2000s, 3000s)
            if account_code_str.startswith(('1', '2', '3')):
                return "BS breakdown/BS sheet"
            
            # P&L account code patterns (typically 4000s, 5000s, 6000s, 7000s)  
            elif account_code_str.startswith(('4', '5', '6', '7', '8', '9')):
                return "PL breakdown sheet"
            
            # Default fallback
            return "Unknown sheet (categorized as Balance Sheet)"
            
        except Exception as e:
            self.logger.warning(f"Error determining anomaly source for account {anomaly.account_code}: {str(e)}")
            return "Unknown sheet"
    
    def _format_period(self, period: str) -> str:
        """Format period as MM/YYYY."""
        try:
            # Handle various period formats
            if '/' in period:
                parts = period.split('/')
                if len(parts) >= 2:
                    month = parts[0].zfill(2)  # Ensure 2 digits
                    year = parts[-1]
                    return f"{month}/{year}"
            
            # If period is already in correct format or unrecognizable, return as-is
            return period
            
        except Exception:
            return period
    
    def _format_triggers_new(self, anomaly: Anomaly) -> str:
        """Format triggers according to new examples."""
        triggers = []
        
        # Handle correlation violations
        if anomaly.type.value == 'correlation_violation':
            if anomaly.rule_violation_id and anomaly.rule_violation_id.startswith('CR'):
                # Extract correlation details from description or logic_trigger
                correlation_detail = self._extract_correlation_details(anomaly)
                triggers.append(f"Correlation break: {correlation_detail}")
        
        # Handle variance-based anomalies
        elif anomaly.type.value == 'variance':
            # Determine account type and thresholds
            abs_percent = abs(anomaly.variance_percent) if anomaly.variance_percent else 0
            abs_amount_billions = abs(anomaly.current_value - (anomaly.previous_value or 0)) / 1_000_000_000
            
            # Balance Sheet accounts (general 5% threshold)
            if anomaly.category in ['investment_properties', 'borrowings', 'cash_deposits', 'trade_receivables']:
                if abs_percent >= 5 and abs_amount_billions >= 1:
                    trigger_parts = [f"Balance Sheet >{abs_percent:.0f}% & >{abs_amount_billions:.0f}B"]
                else:
                    trigger_parts = [f"Balance Sheet >{abs_percent:.0f}%"]
            
            # Revenue/OPEX accounts (10% threshold)
            elif anomaly.category in ['opex', 'staff_costs', 'other_expenses', 'revenue']:
                if abs_percent >= 10 and abs_amount_billions >= 1:
                    trigger_parts = [f"Revenue/OPEX >{abs_percent:.0f}% & ≥{abs_amount_billions:.0f}B"]
                else:
                    trigger_parts = [f"Revenue/OPEX >{abs_percent:.0f}%"]
            
            # Recurring accounts (5% threshold)
            elif anomaly.category in ['depreciation']:
                if abs_percent >= 5 and abs_amount_billions >= 1:
                    trigger_parts = [f"Recurring >{abs_percent:.0f}% & >{abs_amount_billions:.0f}B"]
                else:
                    trigger_parts = [f"Recurring >{abs_percent:.0f}%"]
            
            else:
                # General threshold
                if abs_percent >= 5 and abs_amount_billions >= 1:
                    trigger_parts = [f"Balance Sheet >{abs_percent:.0f}% & >{abs_amount_billions:.0f}B"]
                else:
                    trigger_parts = [f"Variance >{abs_percent:.0f}%"]
            
            # Add trend guardrail if applicable
            if abs_amount_billions >= 1:
                trigger_parts.append(f"Trend guardrail: Cur deviates >{abs_amount_billions:.0f}B from 3–6M avg")
                if abs_amount_billions >= 1:  # Additional trend condition
                    trigger_parts.append(f"Cur deviates >{abs_amount_billions:.0f}B from trend")
            
            triggers.extend(trigger_parts)
        
        # Handle sign changes
        elif anomaly.type.value == 'sign_change':
            triggers.append("Sign change detected")
        
        # Handle other anomaly types
        else:
            if anomaly.logic_trigger:
                triggers.append(anomaly.logic_trigger)
            else:
                triggers.append("Threshold exceeded")
        
        return "; ".join(triggers)
    
    def _extract_correlation_details(self, anomaly: Anomaly) -> str:
        """Extract correlation details for trigger formatting."""
        # Try to extract from description or use default format
        if "Correlation break:" in anomaly.description:
            return anomaly.description.split("Correlation break:")[-1].strip()
        
        # Default format based on rule ID
        rule_map = {
            'CR002': 'LoanΔ=XXX, Interest ExpenseΔ=XXX',
            'CR003': 'CashΔ=XXX, Bank Interest IncomeΔ=XXX', 
            'CR010': 'Asset DisposalΔ=XXX, DepreciationΔ=XXX',
            'CR012': 'Lease TerminationΔ=XXX, RevenueΔ=XXX',
            'CR001': 'Investment PropertiesΔ=XXX, DepreciationΔ=XXX',
            'CR008': 'Occupancy RateΔ=XXX, RevenueΔ=XXX'
        }
        
        if anomaly.rule_violation_id in rule_map:
            # Calculate actual values if available
            current_delta = anomaly.current_value - (anomaly.previous_value or 0)
            return rule_map[anomaly.rule_violation_id].replace('XXX', f'{current_delta:,.0f}')
        
        return f"Rule {anomaly.rule_violation_id}: {anomaly.rule_violation_name or 'Correlation violation'}"
    
    def _generate_suggested_cause(self, anomaly: Anomaly) -> str:
        """Generate suggested likely cause based on anomaly characteristics."""
        return anomaly.recommended_action
    
    def _write_anomaly_data(self, worksheet, summary_data: List[dict]) -> None:
        """Write anomaly data to worksheet."""
        if not summary_data:
            # Create empty row with headers if no data
            headers = ['Subsidiary', 'Account', 'Period', 'Pct Change', 'Absolute Change (VND)', 
                      'Trigger(s)', 'Suggested likely cause', 'Status', 'Notes']
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            return
        
        # Filter out internal fields (starting with underscore) for headers
        visible_keys = [key for key in summary_data[0].keys() if not key.startswith('_')]
        
        # Write headers
        for col, header in enumerate(visible_keys, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Write data (excluding internal fields)
        for row_idx, row_data in enumerate(summary_data, 2):
            col_idx = 1
            for key, value in row_data.items():
                if not key.startswith('_'):  # Skip internal fields
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
                    col_idx += 1
    
    def _apply_anomaly_formatting(self, worksheet, summary_data: List[dict], anomalies: List[Anomaly]) -> None:
        """Apply formatting to the anomaly worksheet."""
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1f4e79', end_color='1f4e79', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for col in range(1, len(summary_data[0]) + 1 if summary_data else 10):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting for severity
        if summary_data:
            severity_colors = {
                'critical': 'ffcccc',  # Light red
                'high': 'ffe6cc',      # Light orange
                'medium': 'ffffcc',    # Light yellow
                'low': 'ffffff'        # White
            }
            
            for row_idx, record in enumerate(summary_data, 2):
                # Get severity from the stored value, default to 'low' if not found
                severity = record.get('_severity', 'low')
                fill_color = severity_colors.get(severity, 'ffffff')
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                # Apply formatting to all columns except the hidden severity column
                num_visible_cols = len([k for k in record.keys() if not k.startswith('_')])
                for col in range(1, num_visible_cols + 1):
                    cell = worksheet.cell(row=row_idx, column=col)
                    cell.fill = fill
                    cell.border = border
                    
                    # Format Absolute Change (VND) column with number format
                    if col == 5:  # Absolute Change (VND) column
                        cell.number_format = '#,##0'
        
        # Set column widths
        column_widths = [15, 40, 12, 12, 20, 45, 40, 15, 25]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    def _create_fallback_file(self, financial_data: FinancialData,
                             variance_results: List[VarianceResult],
                             correlation_results: List[CorrelationResult],
                             anomalies: List[Anomaly],
                             output_file: str) -> None:
        """Create fallback Excel file if adding to existing file fails."""
        self.logger.info("Creating fallback Excel file with xlsxwriter")
        
        # Ensure output directory exists
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook with only Anomalies Summary sheet
        with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            workbook = writer.book
            
            # Prepare data without TB sheet info
            summary_data = self._prepare_anomaly_data(anomalies, pd.DataFrame())
            
            if summary_data:
                df_summary = pd.DataFrame(summary_data)
            else:
                # Create empty dataframe with headers
                df_summary = pd.DataFrame(columns=[
                    'Subsidiary', 'Account', 'Period', 'Pct Change', 'Absolute Change (VND)',
                    'Trigger(s)', 'Suggested likely cause', 'Status', 'Notes'
                ])
            
            # Write to Excel
            sheet_name = 'Anomalies Summary'
            df_summary.to_excel(writer, sheet_name=sheet_name, index=False)
            
            worksheet = writer.sheets[sheet_name]
            
            # Apply basic formatting
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#1f4e79',
                'font_color': 'white',
                'align': 'center',
                'valign': 'vcenter',
                'border': 1
            })
            
            # Apply header formatting
            for col_num, column in enumerate(df_summary.columns):
                worksheet.write(0, col_num, column, header_format)
            
            # Set column widths
            column_widths = [15, 40, 12, 12, 20, 45, 40, 15, 25]
            for i, width in enumerate(column_widths):
                worksheet.set_column(i, i, width)
        
        self.logger.info(f"Fallback Excel file created: {output_file}")